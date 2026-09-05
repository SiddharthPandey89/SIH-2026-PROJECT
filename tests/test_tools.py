"""Combined offline tests for the backend tools layer."""

from __future__ import annotations

import asyncio
import json
import shutil
import uuid
from pathlib import Path
from unittest.mock import AsyncMock

import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[1]

from backend.agent.executer import Executor
from backend.tools import tool_code_exec, tool_file_rw, tool_output_gen, tool_spreadsheet
from backend.tools import tools_doc_search


@pytest.fixture
def tool_root():
    root = PROJECT_ROOT / "sandbox" / "temp_outputs" / f"test_tools_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=False)
    try:
        yield root
    finally:
        shutil.rmtree(root, ignore_errors=True)


def assert_file_result(result, operation=None):
    assert set(result) == {"success", "status", "operation", "path", "content", "error"}
    if operation is not None:
        assert result["operation"] == operation


def assert_spreadsheet_result(result, operation=None):
    assert set(result) == {"success", "status", "operation", "path", "message", "error", "data"}
    if operation is not None:
        assert result["operation"] == operation


def assert_output_result(result):
    assert set(result) == {"success", "status", "format", "path", "message", "error"}


def test_code_exec_success_and_stdout():
    result = tool_code_exec.tool(code="print(6 * 7)")

    assert result["success"] is True
    assert result["status"] == "success"
    assert result["returncode"] == 0
    assert result["stdout"].strip() == "42"
    assert result["stderr"] == ""
    assert result["blocked_patterns"] == []
    json.dumps(result)


def test_code_exec_wrapper_and_structured_contract():
    result = tool_code_exec.tool(code="print('wrapped')", max_output_size=100)

    assert set(result) == {
        "success",
        "status",
        "returncode",
        "stdout",
        "stderr",
        "execution_time",
        "timeout",
        "truncated",
        "blocked_patterns",
    }
    assert result["stdout"].strip() == "wrapped"
    assert result["truncated"] == {"stdout": False, "stderr": False}


def test_code_exec_coerces_non_string_and_empty_code_safely():
    non_string = tool_code_exec.execute_python_code(123)
    empty = tool_code_exec.execute_python_code(None)

    assert non_string["success"] is True
    assert empty["success"] is True
    assert non_string["status"] == "success"
    assert empty["status"] == "success"


@pytest.mark.parametrize("code", ["raise RuntimeError('boom')", "def broken(:\n    pass"])
def test_code_exec_runtime_and_syntax_errors(code):
    result = tool_code_exec.execute_python_code(code)

    assert result["success"] is False
    assert result["status"] == "error"
    assert result["returncode"] is not None
    assert result["stderr"]


def test_code_exec_validation_and_security():
    invalid_timeout = tool_code_exec.execute_python_code("print(1)", timeout=0)
    oversized = tool_code_exec.execute_python_code("x" * 11, max_code_size=10)
    blocked = tool_code_exec.execute_python_code("import subprocess")
    outside_workspace = tool_code_exec.execute_python_code(
        "print(1)", workspace=PROJECT_ROOT / "data"
    )

    assert invalid_timeout["status"] == "rejected"
    assert oversized["status"] == "rejected"
    assert blocked["status"] == "rejected"
    assert blocked["blocked_patterns"]
    assert outside_workspace["status"] == "rejected"


def test_code_exec_timeout_and_output_limit():
    timeout_result = tool_code_exec.execute_python_code(
        "import time; time.sleep(2)", timeout=0.1
    )
    output_result = tool_code_exec.execute_python_code(
        "print('abcdefghij')", max_output_size=4
    )

    assert timeout_result["success"] is False
    assert timeout_result["status"] == "timeout"
    assert output_result["success"] is True
    assert output_result["truncated"]["stdout"] is True
    assert len(output_result["stdout"]) == 4


def test_file_rw_write_append_read_and_contract(tool_root):
    path = tool_root / "notes.txt"

    written = tool_file_rw.tool(operation="write", path=path, content="one")
    appended = tool_file_rw.tool(operation="append", path=path, content=" two")
    read = tool_file_rw.tool(operation="read", path=path)

    assert_file_result(written, "write")
    assert_file_result(appended, "append")
    assert_file_result(read, "read")
    assert written["success"] is True
    assert appended["success"] is True
    assert read["success"] is True
    assert read["content"] == "one two"
    assert path.read_text(encoding="utf-8") == "one two"


def test_file_rw_direct_mode_and_overwrite(tool_root):
    path = tool_root / "direct.txt"

    first = tool_file_rw.tool(path=path, mode="write", content="first")
    second = tool_file_rw.tool(path=path, mode="write", content="second")

    assert first["success"] is True
    assert second["success"] is True
    assert path.read_text(encoding="utf-8") == "second"


def test_file_rw_rejections_and_limits(tool_root):
    missing = tool_file_rw.tool(operation="read", path=tool_root / "missing.txt")
    bad_operation = tool_file_rw.tool(operation="rename", path=tool_root / "x.txt")
    bad_mode = tool_file_rw.tool(path=tool_root / "x.txt", mode="rename")
    outside = tool_file_rw.tool(operation="write", path=PROJECT_ROOT.parent / "outside.txt", content="x")
    traversal = tool_file_rw.tool(
        operation="write",
        path=tool_root / ".." / ".." / ".." / "escape.txt",
        content="x",
    )
    bad_size = tool_file_rw.tool(operation="write", path=tool_root / "x.txt", content="x", max_size=0)
    too_large = tool_file_rw.tool(operation="write", path=tool_root / "x.txt", content="12345", max_size=4)
    bad_encoding = tool_file_rw.tool(operation="write", path=tool_root / "x.txt", content="x", encoding="no-such-encoding")

    assert missing["status"] == "error"
    assert bad_operation["status"] == "rejected"
    assert bad_mode["status"] == "rejected"
    assert outside["status"] == "rejected"
    assert traversal["status"] == "rejected"
    assert bad_size["status"] == "rejected"
    assert too_large["status"] == "rejected"
    assert bad_encoding["status"] == "error"


def test_file_rw_append_size_limit(tool_root):
    path = tool_root / "append.txt"
    assert tool_file_rw.write_file(path, "1234")["success"] is True

    result = tool_file_rw.write_file(path, "56", mode="append", max_size=5)

    assert result["success"] is False
    assert result["status"] == "rejected"
    assert path.read_text(encoding="utf-8") == "1234"


def test_spreadsheet_create_write_read_append_inspect_and_reopen(tool_root):
    path = tool_root / "workbook.xlsx"
    write = tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="Data",
        range="A1:B2",
        data=[["Name", "Value"], ["Alpha", 10]],
        create=True,
        overwrite=True,
        create_sheet=True,
    )
    append = tool_spreadsheet.tool(
        operation="append",
        path=path,
        sheet="Data",
        rows=[["Beta", 20]],
        overwrite=True,
    )
    read = tool_spreadsheet.tool(operation="read", path=path, sheet="Data", range="A1:B3")
    inspect = tool_spreadsheet.tool(operation="inspect", path=path)

    assert_spreadsheet_result(write, "write")
    assert_spreadsheet_result(append, "append")
    assert_spreadsheet_result(read, "read")
    assert_spreadsheet_result(inspect, "inspect")
    assert write["success"] is True
    assert append["data"]["rows_appended"] == 1
    assert read["data"]["values"] == [["Name", "Value"], ["Alpha", 10], ["Beta", 20]]
    assert inspect["data"]["sheets"] == [{"name": "Data", "rows": 3, "columns": 2}]
    assert path.is_file() and path.stat().st_size > 0

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    assert workbook.sheetnames == ["Data"]
    assert workbook["Data"]["B3"].value == 20
    workbook.close()


def test_spreadsheet_multiple_sheets_and_scalar_write(tool_root):
    path = tool_root / "multi.xlsx"
    created = tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="First",
        range="A1",
        data="value",
        create=True,
        overwrite=True,
        create_sheet=True,
    )
    second = tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="Second",
        range="C3",
        data=True,
        overwrite=True,
        create_sheet=True,
    )
    read = tool_spreadsheet.tool(operation="read", path=path, sheet="Second", range="C3")

    assert created["success"] is True
    assert second["success"] is True
    assert read["data"]["values"] == [[True]]
    assert {sheet["name"] for sheet in tool_spreadsheet.tool(operation="inspect", path=path)["data"]["sheets"]} == {"First", "Second"}


@pytest.mark.parametrize(
    ("computation", "expected"),
    [("sum", 60), ("average", 20), ("min", 10), ("max", 30), ("count", 3)],
)
def test_spreadsheet_computations(tool_root, computation, expected):
    path = tool_root / f"{computation}.xlsx"
    created = tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="Numbers",
        range="A1:A3",
        data=[[10], [20], [30]],
        create=True,
        overwrite=True,
        create_sheet=True,
    )
    result = tool_spreadsheet.tool(
        operation="compute",
        path=path,
        sheet="Numbers",
        range="A1:A3",
        computation=computation,
    )

    assert created["success"] is True
    assert result["success"] is True
    assert result["data"]["value"] == expected


def test_spreadsheet_compute_write_and_validation(tool_root):
    path = tool_root / "compute_write.xlsx"
    tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="Numbers",
        range="A1:A2",
        data=[[2], [3]],
        create=True,
        overwrite=True,
        create_sheet=True,
    )
    computed = tool_spreadsheet.tool(
        operation="compute",
        path=path,
        sheet="Numbers",
        range="A1:A2",
        computation="sum",
        destination="B1",
        overwrite=True,
    )
    read_back = tool_spreadsheet.tool(operation="read", path=path, sheet="Numbers", range="B1")
    bad_range = tool_spreadsheet.tool(operation="read", path=path, sheet="Numbers", range="A0")
    bad_matrix = tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="Numbers",
        range="C1:D2",
        data=[[1], [2]],
        overwrite=True,
    )
    bad_type = tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="Numbers",
        range="C1",
        data=object(),
        overwrite=True,
    )
    bad_compute = tool_spreadsheet.tool(
        operation="compute",
        path=path,
        sheet="Numbers",
        range="A1:A2",
        computation="median",
    )

    assert computed["success"] is True
    assert read_back["data"]["values"] == [[5]]
    assert bad_range["status"] == "rejected"
    assert bad_matrix["status"] == "rejected"
    assert bad_type["status"] == "rejected"
    assert bad_compute["status"] == "rejected"


def test_spreadsheet_rejections_and_overwrite(tool_root):
    path = tool_root / "validation.xlsx"
    missing = tool_spreadsheet.tool(operation="inspect", path=path)
    invalid_extension = tool_spreadsheet.tool(operation="inspect", path=tool_root / "bad.xls")
    outside = tool_spreadsheet.tool(operation="inspect", path=PROJECT_ROOT.parent / "outside.xlsx")
    invalid_operation = tool_spreadsheet.tool(operation="delete", path=path)
    missing_sheet = tool_spreadsheet.tool(operation="read", path=path, sheet="Missing", range="A1")
    created = tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="Data",
        range="A1",
        data=1,
        create=True,
        overwrite=True,
        create_sheet=True,
    )
    protected = tool_spreadsheet.tool(
        operation="write",
        path=path,
        sheet="Data",
        range="A1",
        data=2,
    )

    assert missing["status"] == "error"
    assert invalid_extension["status"] == "rejected"
    assert outside["status"] == "rejected"
    assert invalid_operation["status"] == "rejected"
    assert missing_sheet["status"] == "error"
    assert created["success"] is True
    assert protected["status"] == "rejected"


def test_output_generator_all_formats_and_contract(tool_root):
    cases = [
        (
            "docx",
            {"title": "Tool Test", "sections": [{"heading": "Intro", "paragraphs": ["Hello"]}]},
            "result.docx",
        ),
        (
            ".pptx",
            {"title": "Tool Test", "slides": [{"title": "Intro", "content": ["Hello"]}]},
            "result.pptx",
        ),
        (
            "XLSX",
            {"sheets": [{"name": "Data", "headers": ["A"], "rows": [[1]]}]},
            "result.xlsx",
        ),
    ]

    for format_name, content, filename in cases:
        result = tool_output_gen.tool(
            format=format_name,
            content=content,
            path=tool_root / filename,
            overwrite=False,
        )
        assert_output_result(result)
        assert result["success"] is True
        assert result["status"] == "success"
        output_path = Path(result["path"])
        assert output_path == (tool_root / filename).resolve()
        assert output_path.is_file()
        assert output_path.stat().st_size > 0


def test_output_generator_validation_and_overwrite(tool_root):
    path = tool_root / "output.docx"
    content = {"title": "Tool Test"}
    first = tool_output_gen.tool(format="docx", content=content, path=path)
    protected = tool_output_gen.tool(format="docx", content=content, path=path)
    wrong_extension = tool_output_gen.tool(format="docx", content=content, path=tool_root / "wrong.xlsx")
    invalid_format = tool_output_gen.tool(format="pdf", content=content, path=tool_root / "wrong.pdf")
    outside = tool_output_gen.tool(format="docx", content=content, path=PROJECT_ROOT.parent / "outside.docx")
    malformed = tool_output_gen.tool(format="docx", content=[], path=tool_root / "malformed.docx")
    kwargs = tool_output_gen.tool(format="docx", content=content, path=tool_root / "kwargs.docx", unsupported=True)

    assert first["success"] is True
    assert protected["status"] == "rejected"
    assert wrong_extension["status"] == "rejected"
    assert invalid_format["status"] == "rejected"
    assert outside["status"] == "rejected"
    assert malformed["status"] == "rejected"
    assert kwargs["status"] == "rejected"


def test_output_generator_packages_reopen(tool_root):
    docx_path = tool_root / "check.docx"
    pptx_path = tool_root / "check.pptx"
    xlsx_path = tool_root / "check.xlsx"
    assert tool_output_gen.generate_output("docx", {"title": "Check"}, docx_path)["success"] is True
    assert tool_output_gen.generate_output("pptx", {"title": "Check", "slides": [{"title": "One", "content": ["Text"]}]}, pptx_path)["success"] is True
    assert tool_output_gen.generate_output("xlsx", {"sheets": [{"name": "Data", "headers": ["A"], "rows": [[1]]}]}, xlsx_path)["success"] is True

    from docx import Document
    from openpyxl import load_workbook
    from pptx import Presentation

    assert Document(docx_path).paragraphs
    assert len(Presentation(pptx_path).slides) == 2
    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook.sheetnames == ["Data"]
    workbook.close()


def test_document_search_search_health_and_contract(monkeypatch):
    fake = type("FakeRetriever", (), {})()
    fake.retrieve_context = AsyncMock(return_value=[
        {"document_id": "doc-1", "title": "Manual", "snippet": "Pump safety", "score": 0.9}
    ])
    fake.health_check = AsyncMock(return_value=True)
    monkeypatch.setattr(tools_doc_search, "get_retriever", lambda: fake)

    result = asyncio.run(
        tools_doc_search.tool(
            operation="search",
            query=" pump safety ",
            file_id=" manual-1 ",
            top_k=1,
        )
    )
    health = asyncio.run(tools_doc_search.tool(operation="health"))

    assert result["success"] is True
    assert result["data"]["query"] == "pump safety"
    assert result["data"]["file_id"] == "manual-1"
    assert result["data"]["top_k"] == 1
    assert result["data"]["count"] == 1
    assert result["data"]["results"][0]["document_id"] == "doc-1"
    assert health["success"] is True
    fake.retrieve_context.assert_awaited_once_with(query="pump safety", file_id="manual-1", top_k=1)


def test_document_search_empty_results_and_validation(monkeypatch):
    fake = type("FakeRetriever", (), {})()
    fake.retrieve_context = AsyncMock(return_value=[])
    fake.health_check = AsyncMock(return_value=True)
    monkeypatch.setattr(tools_doc_search, "get_retriever", lambda: fake)

    empty = asyncio.run(tools_doc_search.tool(operation="search", query="nothing"))
    calls_after_valid_search = fake.retrieve_context.await_count
    invalid = [
        asyncio.run(tools_doc_search.tool(operation="search", query="")),
        asyncio.run(tools_doc_search.tool(operation="search", query="   ")),
        asyncio.run(tools_doc_search.tool(operation="search", query="x", top_k=0)),
        asyncio.run(tools_doc_search.tool(operation="search", query="x", top_k=-1)),
        asyncio.run(tools_doc_search.tool(operation="search", query="x", top_k=True)),
        asyncio.run(tools_doc_search.tool(operation="search", query="x", top_k="5")),
        asyncio.run(tools_doc_search.tool(operation="search", query="x", file_id="   ")),
        asyncio.run(tools_doc_search.tool(operation="search", query="x", file_id="x" * 257)),
        asyncio.run(tools_doc_search.tool(operation="unknown", query="x")),
        asyncio.run(tools_doc_search.tool(query="x")),
        asyncio.run(tools_doc_search.tool(operation="search", query="x", extra=True)),
    ]

    assert empty["success"] is True
    assert empty["data"]["results"] == []
    assert empty["data"]["count"] == 0
    assert all(item["status"] == "rejected" for item in invalid)
    assert calls_after_valid_search == 1
    assert fake.retrieve_context.await_count == calls_after_valid_search


def test_document_search_malformed_and_failure(monkeypatch):
    fake = type("FakeRetriever", (), {})()
    fake.health_check = AsyncMock(return_value=False)
    fake.retrieve_context = AsyncMock(return_value=[
        {"document_id": "doc", "title": "Title", "snippet": "Text", "score": float("nan")}
    ])
    monkeypatch.setattr(tools_doc_search, "get_retriever", lambda: fake)

    malformed = asyncio.run(tools_doc_search.tool(operation="search", query="x"))
    unhealthy = asyncio.run(tools_doc_search.tool(operation="health"))
    fake.retrieve_context = AsyncMock(side_effect=RuntimeError("backend unavailable"))
    failed = asyncio.run(tools_doc_search.tool(operation="search", query="x"))

    assert malformed["status"] == "error"
    assert unhealthy["status"] == "error"
    assert failed["status"] == "error"
    assert "RuntimeError" in failed["error"]


def test_executor_accepts_sync_and_async_tools(tool_root, monkeypatch):
    fake = type("FakeRetriever", (), {})()
    fake.retrieve_context = AsyncMock(return_value=[])
    fake.health_check = AsyncMock(return_value=True)
    monkeypatch.setattr(tools_doc_search, "get_retriever", lambda: fake)

    executor = Executor(
        tools={
            "file_rw": tool_file_rw.tool,
            "doc_search": tools_doc_search.tool,
        }
    )
    plan = {
        "steps": [
            {
                "id": "step_1",
                "tool": "file_rw",
                "args": {
                    "operation": "write",
                    "path": tool_root / "executor.txt",
                    "content": "from executor",
                },
            },
            {
                "id": "step_2",
                "tool": "doc_search",
                "args": {"operation": "search", "query": "missing"},
            },
        ]
    }

    result = asyncio.run(executor.execute(plan))

    assert result.success is True
    assert len(result.results) == 2
    assert result.results[0].success is True
    assert result.results[1].success is True
    assert (tool_root / "executor.txt").read_text(encoding="utf-8") == "from executor"
