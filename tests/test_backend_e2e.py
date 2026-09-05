"""A-to-Z offline backend validation for the current SIH 2026 repository.

Model inference is intentionally not exercised here: local checkpoints and
model servers are not required for deterministic backend, API, database,
tool, ingestion, and validation coverage.
"""

from __future__ import annotations

import asyncio
import importlib
import json
import shutil
import uuid
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[1]

from backend.agent.executer import (
    Executor,
    InvalidPlanError,
    ToolExecutionError,
    ToolNotFoundError,
)
from backend.agent.planner import InvalidPlanError as PlannerInvalidPlanError
from backend.agent.planner import PlanStep, Planner
from backend.knowledge_base import ingestion, retriever
from backend.model_router import model_registry, task_classifier
from backend.multimodel.ocr_pipeline import OCRPipeline
from backend.multimodel.pdf_parser import PDFParser
from backend.multimodel.vision_pipeline import VisionPipeline
from backend.tools import tool_code_exec, tool_file_rw, tool_output_gen, tool_spreadsheet
from backend.tools import tools_doc_search


@pytest.fixture
def workspace_root():
    root = PROJECT_ROOT / "sandbox" / "temp_outputs" / f"e2e_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=False)
    try:
        yield root
    finally:
        shutil.rmtree(root, ignore_errors=True)


def run(coro):
    return asyncio.run(coro)


def assert_result_keys(result, keys):
    assert set(result) == set(keys)
    json.dumps(result)


@pytest.fixture
def api_client(monkeypatch, workspace_root):
    """Create a client without entering the production lifespan/database."""
    pytest.importorskip("sqlalchemy", reason="SQLAlchemy is required for API/database integration tests.")
    pytest.importorskip("fastapi", reason="FastAPI is required for API integration tests.")
    from fastapi.testclient import TestClient
    from backend.api import routes_chat, routes_knowledge, routes_agent, routes_upload
    from backend.database.init_db import get_db
    from backend.model_router.router import get_model_router
    from backend.knowledge_base.retriever import get_retriever
    from backend.agent.executer import get_executor
    from backend.agent.planner import get_planner

    class FakeModel:
        async def generate(self, message, task_type, history, context_chunks=None):
            return {"answer": f"offline test answer: {message}", "model": "test-model"}

        async def health_check(self):
            return True

    class FakeRetriever:
        async def retrieve_context(self, query, file_id=None, top_k=5):
            return [{
                "document_id": "doc-e2e",
                "title": "E2E Document",
                "snippet": "Local test context",
                "score": 0.9,
            }]

        async def health_check(self):
            return True

    class FakePlanner:
        async def create_plan(self, *args, **kwargs):
            return {"goal": "test", "task_type": "code", "steps": []}

    class FakeDatabase:
        pass

    async def override_db():
        yield FakeDatabase()

    async def fake_create_conversation(db, title=None):
        return SimpleNamespace(id="e2e-conversation")

    async def fake_get_conversation(db, conversation_id):
        if conversation_id == "missing":
            return None
        return SimpleNamespace(id=conversation_id)

    async def fake_history(db, conversation_id):
        return []

    async def fake_add_message(*args, **kwargs):
        return None

    async def fake_ping(db):
        return True

    monkeypatch.setattr(routes_chat.crud, "create_conversation", fake_create_conversation)
    monkeypatch.setattr(routes_chat.crud, "get_conversation", fake_get_conversation)
    monkeypatch.setattr(routes_chat.crud, "get_conversation_history", fake_history)
    monkeypatch.setattr(routes_chat.crud, "add_message", fake_add_message)
    monkeypatch.setattr(routes_chat.crud, "ping", fake_ping)
    monkeypatch.setattr(routes_agent.crud, "create_conversation", fake_create_conversation)
    monkeypatch.setattr(routes_agent.crud, "get_conversation", fake_get_conversation)
    monkeypatch.setattr(routes_agent.crud, "get_conversation_history", fake_history)
    monkeypatch.setattr(routes_agent.crud, "add_message", fake_add_message)
    monkeypatch.setattr(routes_chat, "classify_task", AsyncMock(return_value="chat"))
    monkeypatch.setattr(routes_upload, "UPLOAD_ROOT", workspace_root / "uploads")

    from backend.main import app

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_model_router] = lambda: FakeModel()
    app.dependency_overrides[routes_chat.get_model_router] = lambda: FakeModel()
    app.dependency_overrides[get_retriever] = lambda: FakeRetriever()
    app.dependency_overrides[routes_chat.get_retriever] = lambda: FakeRetriever()
    app.dependency_overrides[routes_knowledge.get_retriever] = lambda: FakeRetriever()
    app.dependency_overrides[routes_agent.get_retriever] = lambda: FakeRetriever()
    app.dependency_overrides[get_planner] = lambda: FakePlanner()
    app.dependency_overrides[get_executor] = lambda: Executor()

    client = TestClient(app)
    try:
        yield client
    finally:
        client.close()
        app.dependency_overrides.clear()


def test_backend_imports_and_public_app():
    modules = [
        "backend.config",
        "backend.main",
        "backend.api.schemas",
        "backend.api.routes_chat",
        "backend.api.routes_agent",
        "backend.api.routes_knowledge",
        "backend.api.routes_upload",
        "backend.agent.executer",
        "backend.agent.planner",
        "backend.agent.memory",
        "backend.database.crud",
        "backend.database.db_models",
        "backend.database.init_db",
        "backend.knowledge_base.ingestion",
        "backend.knowledge_base.retriever",
        "backend.model_router.model_registry",
        "backend.model_router.router",
        "backend.model_router.task_classifier",
        "backend.multimodel.pdf_parser",
        "backend.multimodel.ocr_pipeline",
        "backend.multimodel.vision_pipeline",
        "backend.tools.tool_code_exec",
        "backend.tools.tool_file_rw",
        "backend.tools.tool_spreadsheet",
        "backend.tools.tool_output_gen",
        "backend.tools.tools_doc_search",
    ]
    for name in modules:
        try:
            imported = importlib.import_module(name)
        except ModuleNotFoundError as exc:
            if exc.name and not exc.name.startswith("backend"):
                pytest.skip(f"Optional dependency unavailable for {name}: {exc.name}")
            raise
        assert imported is not None

    from fastapi import FastAPI
    from backend.main import app

    assert isinstance(app, FastAPI)
    assert app.openapi()["info"]["title"] == "Sovereign AI Workbench"


def test_api_root_health_openapi_and_route_registration(api_client):
    assert api_client.get("/").json()["status"] == "running"
    assert api_client.get("/health").json() == {"status": "ok"}
    schema = api_client.get("/openapi.json")
    assert schema.status_code == 200
    paths = schema.json()["paths"]
    for path in ("/api/chat", "/api/upload", "/api/knowledge/search", "/api/agent/run"):
        assert path in paths


def test_chat_api_validation_and_mocked_contract(api_client):
    assert api_client.post("/api/chat", json={}).status_code == 422
    assert api_client.post("/api/chat", json={"message": 123}).status_code == 422
    assert api_client.post("/api/chat", json={"message": ""}).status_code == 400
    assert api_client.post("/api/chat", json={"message": "ok", "extra": True}).status_code == 422

    response = api_client.post("/api/chat", json={"message": "hello"})
    assert response.status_code == 200
    body = response.json()
    assert {"answer", "model", "task_type", "conversation_id", "sources", "created_at"} <= body.keys()
    assert body["model"] == "test-model"


def test_chat_and_knowledge_health_and_search(api_client):
    chat_health = api_client.get("/api/chat/health")
    knowledge_health = api_client.get("/api/knowledge/health")
    search = api_client.get("/api/knowledge/search", params={"query": "pump", "top_k": 1})
    documents = api_client.get("/api/knowledge/documents")

    assert chat_health.status_code == 200
    assert chat_health.json()["status"] == "ok"
    assert knowledge_health.status_code == 200
    assert knowledge_health.json()["knowledge_base_ready"] is True
    assert search.status_code == 200
    assert search.json()["count"] == 1
    assert search.json()["results"][0]["document_id"] == "doc-e2e"
    assert documents.status_code == 200
    assert documents.json()["supported"] is False


def test_knowledge_api_validation(api_client):
    assert api_client.get("/api/knowledge/search").status_code == 422
    assert api_client.get("/api/knowledge/search", params={"query": "x", "top_k": 0}).status_code == 422
    assert api_client.get("/api/knowledge/search", params={"query": "x", "top_k": 51}).status_code == 422


def test_agent_api_validation_and_unknown_status(api_client):
    invalid = api_client.post("/api/agent/run", json={})
    unknown = api_client.get("/api/agent/status/does-not-exist")

    assert invalid.status_code == 422
    assert unknown.status_code == 404
    assert "detail" in unknown.json()


def test_upload_api_storage_flow(api_client, workspace_root):
    upload = api_client.post(
        "/api/upload",
        files={"file": ("../safe.txt", b"industrial test text", "text/plain")},
    )
    assert upload.status_code == 201
    metadata = upload.json()
    assert {"file_id", "filename", "content_type", "category", "size_bytes", "uploaded_at"} <= metadata.keys()
    assert metadata["category"] == "document"
    assert ".." not in metadata["filename"]

    file_id = metadata["file_id"]
    listed = api_client.get("/api/upload")
    fetched = api_client.get(f"/api/upload/{file_id}")
    deleted = api_client.delete(f"/api/upload/{file_id}")
    missing = api_client.get(f"/api/upload/{file_id}")

    assert listed.status_code == 200
    assert listed.json()["count"] == 1
    assert fetched.status_code == 200
    assert deleted.status_code == 200
    assert deleted.json()["deleted"] is True
    assert missing.status_code == 404
    assert not list((workspace_root / "uploads").rglob(f"{file_id}*"))


def test_upload_api_rejects_invalid_files(api_client):
    assert api_client.post("/api/upload", files={"file": ("bad.exe", b"x", "application/octet-stream")}).status_code == 415
    assert api_client.post("/api/upload", files={"file": ("empty.txt", b"", "text/plain")}).status_code == 400
    assert api_client.get("/api/upload/not-a-valid-id").status_code == 404


def test_executor_sync_async_references_and_failures(workspace_root):
    async def async_tool(value):
        return {"value": value}

    def sync_tool(path, content):
        return tool_file_rw.tool(operation="write", path=path, content=content)

    executor = Executor({"async": async_tool, "sync": sync_tool})
    plan = {
        "steps": [
            {"id": "one", "tool": "async", "args": {"value": 7}},
            {"id": "two", "tool": "sync", "args": {"path": workspace_root / "x.txt", "content": "$one"}},
        ]
    }
    result = run(executor.execute(plan))

    assert result.success is True
    assert len(result.results) == 2
    assert (workspace_root / "x.txt").exists()
    assert result.outputs["one"] == {"value": 7}

    with pytest.raises(InvalidPlanError):
        run(executor.execute({"steps": [{"id": "x", "tool": "sync", "args": [], "continue_on_error": False}]}))
    with pytest.raises(ToolNotFoundError):
        run(executor._invoke_tool("missing", {}))

    failing = Executor({"bad": lambda: (_ for _ in ()).throw(RuntimeError("boom"))})
    failed = run(failing.execute({"steps": [{"id": "x", "tool": "bad", "args": {}}]}))
    assert failed.success is False
    assert failed.results[0].success is False


def test_database_crud_isolated_sqlite():
    pytest.importorskip("sqlalchemy", reason="SQLAlchemy is not installed in this environment.")
    pytest.importorskip("aiosqlite", reason="aiosqlite is not installed in this environment.")
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
    from backend.database import crud
    from backend.database.db_models import Base

    async def exercise():
        engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        try:
            async with engine.begin() as connection:
                await connection.run_sync(Base.metadata.create_all)
            sessions = async_sessionmaker(engine, expire_on_commit=False)
            async with sessions() as session:
                conversation = await crud.create_conversation(session, title="E2E")
                assert conversation.id
                await crud.add_message(session, conversation.id, "user", "hello")
                await crud.add_message(
                    session,
                    conversation.id,
                    "assistant",
                    "world",
                    metadata={"model": "test", "task_type": "chat"},
                )
                history = await crud.get_conversation_history(session, conversation.id)
                assert history == [
                    {"role": "user", "content": "hello"},
                    {"role": "assistant", "content": "world"},
                ]
                assert await crud.get_conversation(session, "missing") is None
                assert await crud.ping(session) is True
                with pytest.raises(crud.InvalidRoleError):
                    await crud.add_message(session, conversation.id, "invalid", "x")
        finally:
            await engine.dispose()

    run(exercise())


def test_planner_parser_and_serialization():
    step = PlanStep(id="step_1", tool="file_rw", args={"operation": "read"})
    assert step.to_dict()["id"] == "step_1"
    parsed = Planner._parse_plan(
        "```json\n{\"goal\": \"read\", \"steps\": [{\"id\": \"step_1\", \"tool\": \"file_rw\", \"args\": {}}]}\n```",
        goal="fallback",
        task_type="code",
    )
    assert parsed["goal"] == "read"
    assert parsed["steps"][0]["tool"] == "file_rw"
    with pytest.raises(PlannerInvalidPlanError):
        Planner._parse_plan("{\"steps\": [{\"id\": \"x\", \"args\": {}}]}", "g", "chat")


def test_task_classifier_is_deterministic_and_local():
    result = run(task_classifier.classify("calculate the total in this spreadsheet"))
    assert result.task_type == "spreadsheet"
    assert 0 <= result.confidence <= 1
    assert isinstance(result.scores, dict)
    assert run(task_classifier.classify_task("write a Python function")) == "code"
    assert run(task_classifier.classify_task("hello there")) == "chat"


def test_model_registry_metadata_and_local_endpoints():
    registry = model_registry.ModelRegistry()
    config = model_registry.ModelConfig(
        model_id="test-model",
        display_name="Test Model",
        backend=model_registry.ModelBackend.TRANSFORMERS,
        endpoint="local://test",
        backend_model_name="test",
        supported_tasks=["chat"],
        context_window=128,
        priority=1,
    )
    registry.register(config)
    assert registry.get("test-model").model_id == "test-model"
    assert registry.get_default_model_for_task("chat").model_id == "test-model"
    assert registry.all_supported_task_types() == ["chat"]
    with pytest.raises(model_registry.DuplicateModelError):
        registry.register(config)
    assert all("openai.com" not in item.endpoint for item in model_registry.get_model_registry().list_models())


def test_model_router_model_independent_failure_and_message_building():
    from backend.model_router.router import ModelRouter, NoModelAvailableError, _build_messages

    empty = model_registry.ModelRegistry()
    router = ModelRouter(registry=empty, adapters={})
    with pytest.raises(NoModelAvailableError):
        run(router.generate("hello", "chat"))
    messages = _build_messages("hello", [{"role": "user", "content": "prior"}], [{"title": "Doc", "snippet": "Context"}])
    assert messages[-1] == {"role": "user", "content": "hello"}
    assert any(message["role"] == "system" for message in messages)


def test_file_tool_and_code_tool_security_boundaries(workspace_root):
    outside = tool_file_rw.tool(operation="write", path=PROJECT_ROOT.parent / "escape.txt", content="x")
    traversal = tool_file_rw.tool(operation="write", path=workspace_root / ".." / ".." / ".." / "escape.txt", content="x")
    blocked = tool_code_exec.execute_python_code("import subprocess")
    invalid_workspace = tool_code_exec.execute_python_code("print(1)", workspace=PROJECT_ROOT / "data")

    assert outside["status"] == "rejected"
    assert traversal["status"] == "rejected"
    assert blocked["status"] == "rejected"
    assert invalid_workspace["status"] == "rejected"


def test_spreadsheet_end_to_end(workspace_root):
    path = workspace_root / "e2e.xlsx"
    created = tool_spreadsheet.tool(
        operation="write", path=path, sheet="Data", range="A1:B2",
        data=[["Item", "Amount"], ["A", 10]], create=True,
        create_sheet=True, overwrite=True,
    )
    appended = tool_spreadsheet.tool(
        operation="append", path=path, sheet="Data", rows=[["B", 20]], overwrite=True,
    )
    computed = tool_spreadsheet.tool(
        operation="compute", path=path, sheet="Data", range="B2:B3", computation="sum",
    )
    read = tool_spreadsheet.tool(operation="read", path=path, sheet="Data", range="A1:B3")

    assert created["success"] and appended["success"] and computed["success"]
    assert computed["data"]["value"] == 30
    assert read["data"]["values"] == [["Item", "Amount"], ["A", 10], ["B", 20]]
    assert path.stat().st_size > 0

    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["Data"]
    workbook.close()


def test_output_generation_end_to_end(workspace_root):
    outputs = [
        ("docx", {"title": "E2E"}, workspace_root / "e2e.docx"),
        ("pptx", {"title": "E2E", "slides": [{"title": "One", "content": ["Text"]}]}, workspace_root / "e2e.pptx"),
        ("xlsx", {"sheets": [{"name": "Data", "headers": ["A"], "rows": [[1]]}]}, workspace_root / "e2e.xlsx"),
    ]
    for format_name, content, path in outputs:
        result = tool_output_gen.generate_output(format_name, content, path)
        assert set(result) == {"success", "status", "format", "path", "message", "error"}
        assert result["success"] is True
        assert path.is_file() and path.stat().st_size > 0


def test_document_search_adapter_with_fake_retriever(monkeypatch):
    fake = SimpleNamespace(
        retrieve_context=AsyncMock(return_value=[
            {"document_id": "d", "title": "T", "snippet": "S", "score": 0.5}
        ]),
        health_check=AsyncMock(return_value=True),
    )
    monkeypatch.setattr(tools_doc_search, "get_retriever", lambda: fake)
    result = run(tools_doc_search.tool(operation="search", query=" test ", top_k=1))
    health = run(tools_doc_search.tool(operation="health"))
    assert result["success"] is True
    assert result["data"]["query"] == "test"
    assert result["data"]["count"] == 1
    assert health["data"]["healthy"] is True

    invalid = run(tools_doc_search.tool(operation="search", query="x", top_k=True))
    assert invalid["status"] == "rejected"


def test_ingestion_and_retriever_local_fallback(workspace_root, monkeypatch):
    document = workspace_root / "manual.txt"
    document.write_text("Pump maintenance and lockout tagout procedure.", encoding="utf-8")
    monkeypatch.setattr(ingestion, "_DOCUMENTS_ROOT", workspace_root.resolve())
    monkeypatch.setattr(ingestion, "_UPLOADS_ROOT", (workspace_root / "uploads").resolve())

    result = ingestion.ingest_file(document, chunk_size=20, chunk_overlap=5)
    assert result.success is True
    assert result.status == ingestion.STATUS_SUCCESS
    assert result.chunk_count >= 1
    assert result.file_id
    assert ingestion.chunk_text("abcdef", chunk_size=3, chunk_overlap=1)
    with pytest.raises(ValueError):
        ingestion.chunk_text("x", chunk_size=0)


def test_pdf_parser_and_multimodal_validation(workspace_root):
    outside = PDFParser().parse(PROJECT_ROOT.parent / "outside.pdf")
    wrong_type = PDFParser().parse(workspace_root / "file.txt")
    missing_pdf = PDFParser().parse(workspace_root / "missing.pdf")
    ocr_outside = OCRPipeline().process(PROJECT_ROOT.parent / "outside.pdf")
    vision_outside = VisionPipeline().caption(PROJECT_ROOT.parent / "outside.png")

    assert outside.status == "path_denied"
    assert wrong_type.status in {"path_denied", "unsupported_type"}
    assert missing_pdf.status == "path_denied"
    assert ocr_outside.status == "path_denied"
    assert vision_outside.status == "path_denied"


def test_model_dependent_paths_are_explicitly_skipped():
    """Inference is not falsely reported when local checkpoints are absent."""
    required = [
        PROJECT_ROOT / "models" / "llm",
        PROJECT_ROOT / "models" / "embedding",
        PROJECT_ROOT / "models" / "vision",
    ]
    if all(path.exists() and any(path.iterdir()) for path in required):
        pytest.skip("Model-dependent execution is outside this deterministic suite.")
    pytest.skip("Model weights are not installed; model-dependent execution is skipped.")


def test_fastapi_route_method_restrictions(api_client):
    assert api_client.get("/api/chat").status_code == 405
    assert api_client.post("/health").status_code == 405
    assert api_client.get("/api/upload/not-a-valid-id").status_code == 404


def test_offline_configuration_and_result_serialization():
    from backend import config

    assert config.PROJECT_ROOT == PROJECT_ROOT
    assert isinstance(config.MISTRAL_MODEL_PATH, str)
    assert "http://" not in json.dumps(tool_code_exec.execute_python_code("print(1)"))
    assert json.dumps(tool_spreadsheet.tool(operation="inspect", path=PROJECT_ROOT / "data" / "missing.xlsx"))
