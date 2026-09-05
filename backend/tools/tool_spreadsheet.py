"""Local, workspace-bound spreadsheet tool for XLSX workbooks.

The public ``tool(**kwargs)`` entry point is intended for the agent executor.
All workbook processing is local and uses openpyxl directly. This module
handles workbook inspection, bounded reads, conservative updates/appends, and
small deterministic computations; formatted workbook generation remains the
responsibility of ``backend.output_gen.xlsx_writer``.
"""

from __future__ import annotations

import logging
import math
import os
import re
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

from backend.tools.tool_file_rw import ALLOWED_ROOTS

__all__ = ["spreadsheet_tool", "tool"]

logger = logging.getLogger(__name__)

_PROJECT_FILE_SIZE_LIMIT = 50_000_000
_MAX_ROWS = 100_000
_MAX_COLUMNS = 100
_MAX_SHEETS = 100
_MAX_TEXT_LENGTH = 1_000_000

_CELL_PATTERN = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*$")
_SHEET_INVALID_CHARS = set("[]:*?/\\")
_OPERATIONS = {"inspect", "read", "write", "append", "compute"}
_COMPUTATIONS = {"sum", "average", "min", "max", "count"}
_MISSING = object()


def _result(
    success: bool,
    status: str,
    operation: Optional[str],
    path: Optional[str],
    message: Optional[str],
    error: Optional[str] = None,
    data: Any = None,
) -> Dict[str, Any]:
    """Return the stable JSON-compatible spreadsheet-tool result."""
    return {
        "success": success,
        "status": status,
        "operation": operation,
        "path": path,
        "message": message,
        "error": error,
        "data": data,
    }


def _reject(operation: Optional[str], message: str) -> Dict[str, Any]:
    return _result(
        False,
        "rejected",
        operation,
        None,
        message,
    )


def _error(
    operation: str,
    path: Optional[str],
    message: str,
    detail: Optional[str] = None,
) -> Dict[str, Any]:
    return _result(
        False,
        "error",
        operation,
        path,
        message,
        detail or message,
    )


def _resolve_path(path_value: Any) -> Optional[Path]:
    if not isinstance(path_value, (str, Path)):
        return None

    try:
        candidate = Path(path_value).expanduser().resolve()
    except (TypeError, ValueError, OSError):
        return None

    for root in ALLOWED_ROOTS:
        try:
            candidate.relative_to(root)
            return candidate
        except ValueError:
            continue

    return None


def _validate_xlsx_path(
    path_value: Any,
) -> Tuple[Optional[Path], Optional[str]]:
    resolved = _resolve_path(path_value)

    if resolved is None:
        return (
            None,
            "Path must be inside the project's data/ or sandbox/ directory.",
        )

    if resolved.suffix.lower() != ".xlsx":
        return None, "Spreadsheet path must use the .xlsx extension."

    if not resolved.name:
        return None, "Path must include a workbook filename."

    return resolved, None


def _validate_sheet_name(name: Any) -> Optional[str]:
    if not isinstance(name, str) or not name.strip():
        return "Sheet name must be a non-empty string."

    if len(name) > 31:
        return "Sheet name must not exceed 31 characters."

    if any(character in _SHEET_INVALID_CHARS for character in name):
        return "Sheet name contains an invalid Excel character."

    if name.endswith((".", " ")):
        return "Sheet name must not end with a period or space."

    return None


def _validate_scalar(value: Any, label: str) -> Optional[str]:
    if value is None or isinstance(value, (str, int, bool, date, datetime)):
        if isinstance(value, str) and len(value) > _MAX_TEXT_LENGTH:
            return f"{label} exceeds the maximum text length."
        return None

    if isinstance(value, float):
        if not math.isfinite(value):
            return f"{label} must be finite."
        return None

    return (
        f"Unsupported value type for {label}: "
        f"{type(value).__name__}."
    )


def _parse_reference(
    reference: Any,
) -> Tuple[
    Optional[Tuple[int, int, int, int]],
    Optional[str],
]:
    if not isinstance(reference, str) or not reference.strip():
        return None, "Cell or range reference must be a non-empty string."

    value = reference.strip().upper()

    if ":" not in value and not _CELL_PATTERN.fullmatch(value):
        return None, "Invalid cell reference. Use a form such as A1 or A1:C3."

    if ":" in value:
        start, end = value.split(":", 1)

        if (
            not _CELL_PATTERN.fullmatch(start)
            or not _CELL_PATTERN.fullmatch(end)
        ):
            return (
                None,
                "Invalid range reference. Use a form such as A1:C3.",
            )

        value = f"{start}:{end}"

    try:
        min_col, min_row, max_col, max_row = range_boundaries(value)
    except ValueError:
        return None, "Invalid cell or range reference."

    if (
        min_row < 1
        or max_row > _MAX_ROWS
        or min_col < 1
        or max_col > _MAX_COLUMNS
    ):
        return None, "Cell or range exceeds the supported worksheet limits."

    if min_row > max_row or min_col > max_col:
        return None, "Range start must not be after its end."

    return (min_row, min_col, max_row, max_col), None


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, (str, int, float, bool)) or value is None:
        return value

    return str(value)


def _validate_matrix(
    data: Any,
    rows: int,
    columns: int,
) -> Optional[str]:
    if not isinstance(data, (list, tuple)) or len(data) != rows:
        return f"Data must contain exactly {rows} row(s)."

    for row_index, row in enumerate(data):
        if not isinstance(row, (list, tuple)) or len(row) != columns:
            return (
                f"Row {row_index} must contain exactly "
                f"{columns} cell(s)."
            )

        for column_index, value in enumerate(row):
            error = _validate_scalar(
                value,
                f"Cell ({row_index}, {column_index})",
            )
            if error:
                return error

    return None


def _load_existing(
    path: Path,
    operation: str,
) -> Tuple[Optional[Any], Optional[Dict[str, Any]]]:
    if not path.exists():
        return (
            None,
            _error(
                operation,
                str(path),
                "Workbook does not exist.",
            ),
        )

    if not path.is_file():
        return (
            None,
            _error(
                operation,
                str(path),
                "Path is not a regular file.",
            ),
        )

    try:
        if path.stat().st_size > _PROJECT_FILE_SIZE_LIMIT:
            return (
                None,
                _error(
                    operation,
                    str(path),
                    "Workbook exceeds the maximum supported file size.",
                ),
            )

        return (
            load_workbook(
                path,
                read_only=False,
                data_only=False,
            ),
            None,
        )

    except (InvalidFileException, OSError, ValueError) as exc:
        return (
            None,
            _error(
                operation,
                str(path),
                "Unable to open the XLSX workbook.",
                type(exc).__name__,
            ),
        )


def _sheet_or_error(
    workbook: Any,
    sheet_name: Any,
    operation: str,
    path: str,
) -> Tuple[Optional[Any], Optional[Dict[str, Any]]]:
    error = _validate_sheet_name(sheet_name)

    if error:
        return None, _reject(operation, error)

    if sheet_name not in workbook.sheetnames:
        return (
            None,
            _error(
                operation,
                path,
                f"Worksheet '{sheet_name}' does not exist.",
            ),
        )

    return workbook[sheet_name], None


def _range_values(
    worksheet: Any,
    bounds: Tuple[int, int, int, int],
) -> List[List[Any]]:
    min_row, min_col, max_row, max_col = bounds

    return [
        [
            _json_value(
                worksheet.cell(
                    row=row,
                    column=column,
                ).value
            )
            for column in range(min_col, max_col + 1)
        ]
        for row in range(min_row, max_row + 1)
    ]


def _verify_saved_workbook(
    path: Path,
    expected: Callable[[Any], Optional[str]],
) -> Optional[str]:
    try:
        if not path.is_file() or path.stat().st_size == 0:
            return "Saved workbook is missing or empty."

        workbook = load_workbook(
            path,
            read_only=False,
            data_only=False,
        )

        try:
            return expected(workbook)
        finally:
            workbook.close()

    except (InvalidFileException, OSError, ValueError) as exc:
        return (
            "Saved workbook failed verification: "
            f"{type(exc).__name__}."
        )


def _atomic_save(
    workbook: Any,
    path: Path,
    expected: Callable[[Any], Optional[str]],
) -> Optional[str]:
    path.parent.mkdir(parents=True, exist_ok=True)

    temp_fd, temp_name = tempfile.mkstemp(
        prefix=".tmp_",
        suffix=".xlsx",
        dir=str(path.parent),
    )

    os.close(temp_fd)
    temporary = Path(temp_name)

    try:
        workbook.save(temporary)

        verification_error = _verify_saved_workbook(
            temporary,
            expected,
        )

        if verification_error:
            return verification_error

        temporary.replace(path)

        return _verify_saved_workbook(
            path,
            expected,
        )

    except (OSError, ValueError) as exc:
        return (
            "Unable to save workbook safely: "
            f"{type(exc).__name__}."
        )

    finally:
        try:
            temporary.unlink(missing_ok=True)
        except OSError:
            pass


def _inspect(path: Path) -> Dict[str, Any]:
    workbook, failure = _load_existing(
        path,
        "inspect",
    )

    if failure:
        return failure

    try:
        sheets = [
            {
                "name": name,
                "rows": workbook[name].max_row,
                "columns": workbook[name].max_column,
            }
            for name in workbook.sheetnames
        ]

        return _result(
            True,
            "success",
            "inspect",
            str(path),
            "Workbook inspected successfully.",
            data={"sheets": sheets},
        )

    finally:
        workbook.close()


def _read(
    path: Path,
    sheet_name: Any,
    reference: Any,
) -> Dict[str, Any]:
    workbook, failure = _load_existing(
        path,
        "read",
    )

    if failure:
        return failure

    try:
        worksheet, failure = _sheet_or_error(
            workbook,
            sheet_name,
            "read",
            str(path),
        )

        if failure:
            return failure

        if reference is None:
            if (
                worksheet.max_row > _MAX_ROWS
                or worksheet.max_column > _MAX_COLUMNS
            ):
                return _error(
                    "read",
                    str(path),
                    "Worksheet exceeds the supported read dimensions.",
                )

            bounds = (
                1,
                1,
                max(worksheet.max_row, 1),
                max(worksheet.max_column, 1),
            )

        else:
            bounds, error = _parse_reference(reference)

            if error:
                return _reject("read", error)

        values = _range_values(
            worksheet,
            bounds,
        )

        return _result(
            True,
            "success",
            "read",
            str(path),
            "Worksheet data read successfully.",
            data={
                "sheet": sheet_name,
                "range": (
                    reference
                    or f"A1:{get_column_letter(bounds[3])}{bounds[2]}"
                ),
                "values": values,
            },
        )

    finally:
        workbook.close()


def _prepare_mutation(
    path: Path,
    operation: str,
    overwrite: bool,
    create: bool,
) -> Tuple[Optional[Any], Optional[Dict[str, Any]]]:
    if not isinstance(overwrite, bool) or not isinstance(create, bool):
        return (
            None,
            _reject(
                operation,
                "overwrite and create must be boolean values.",
            ),
        )

    if path.exists():
        if path.is_dir():
            return (
                None,
                _error(
                    operation,
                    str(path),
                    "Path is a directory, not a workbook file.",
                ),
            )

        if not overwrite:
            return (
                None,
                _reject(
                    operation,
                    "Existing workbook requires overwrite=True for modifications.",
                ),
            )

        return _load_existing(
            path,
            operation,
        )

    if not create:
        return (
            None,
            _reject(
                operation,
                "Workbook does not exist; use create=True to create it.",
            ),
        )

    workbook = Workbook()

    return workbook, None


def _write(
    path: Path,
    sheet_name: Any,
    reference: Any,
    data: Any,
    overwrite: bool,
    create: bool,
    create_sheet: bool,
) -> Dict[str, Any]:
    workbook, failure = _prepare_mutation(
        path,
        "write",
        overwrite,
        create,
    )

    if failure:
        return failure

    try:
        if not isinstance(create_sheet, bool):
            return _reject(
                "write",
                "create_sheet must be a boolean value.",
            )

        sheet_error = _validate_sheet_name(sheet_name)

        if sheet_error:
            return _reject("write", sheet_error)

        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

        elif create_sheet or create:
            if len(workbook.sheetnames) >= _MAX_SHEETS:
                return _reject(
                    "write",
                    "Workbook has reached the worksheet limit.",
                )

            if (
                workbook.sheetnames == ["Sheet"]
                and workbook["Sheet"].max_row == 1
                and workbook["Sheet"].max_column == 1
                and workbook["Sheet"]["A1"].value is None
            ):
                worksheet = workbook["Sheet"]
                worksheet.title = sheet_name
            else:
                worksheet = workbook.create_sheet(
                    title=sheet_name,
                )

        else:
            return _error(
                "write",
                str(path),
                f"Worksheet '{sheet_name}' does not exist.",
            )

        bounds, error = _parse_reference(reference)

        if error:
            return _reject("write", error)

        min_row, min_col, max_row, max_col = bounds

        height = max_row - min_row + 1
        width = max_col - min_col + 1

        if height == 1 and width == 1:
            error = _validate_scalar(
                data,
                "Cell value",
            )

            if error:
                return _reject("write", error)

            worksheet.cell(
                min_row,
                min_col,
            ).value = data

            expected_values = [[_json_value(data)]]

        else:
            error = _validate_matrix(
                data,
                height,
                width,
            )

            if error:
                return _reject("write", error)

            for row_offset, row in enumerate(data):
                for column_offset, value in enumerate(row):
                    worksheet.cell(
                        min_row + row_offset,
                        min_col + column_offset,
                    ).value = value

            expected_values = [
                [_json_value(value) for value in row]
                for row in data
            ]

        def expected(verified: Any) -> Optional[str]:
            if sheet_name not in verified.sheetnames:
                return "Expected worksheet is missing after save."

            actual_values = _range_values(
                verified[sheet_name],
                bounds,
            )

            if actual_values != expected_values:
                return (
                    "Written range does not match the requested data "
                    "after save."
                )

            return None

        save_error = _atomic_save(
            workbook,
            path,
            expected,
        )

        if save_error:
            return _error(
                "write",
                str(path),
                "Workbook write failed verification.",
                save_error,
            )

        return _result(
            True,
            "success",
            "write",
            str(path),
            "Spreadsheet data written successfully.",
            data={
                "sheet": sheet_name,
                "range": reference,
            },
        )

    finally:
        workbook.close()


def _append(
    path: Path,
    sheet_name: Any,
    rows: Any,
    overwrite: bool,
    create: bool,
    create_sheet: bool,
) -> Dict[str, Any]:
    workbook, failure = _prepare_mutation(
        path,
        "append",
        overwrite,
        create,
    )

    if failure:
        return failure

    try:
        if not isinstance(create_sheet, bool):
            return _reject(
                "append",
                "create_sheet must be a boolean value.",
            )

        sheet_error = _validate_sheet_name(sheet_name)

        if sheet_error:
            return _reject("append", sheet_error)

        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

        elif create_sheet or create:
            if len(workbook.sheetnames) >= _MAX_SHEETS:
                return _reject(
                    "append",
                    "Workbook has reached the worksheet limit.",
                )

            if (
                workbook.sheetnames == ["Sheet"]
                and workbook["Sheet"].max_row == 1
                and workbook["Sheet"].max_column == 1
                and workbook["Sheet"]["A1"].value is None
            ):
                worksheet = workbook["Sheet"]
                worksheet.title = sheet_name
            else:
                worksheet = workbook.create_sheet(
                    title=sheet_name,
                )

        else:
            return _error(
                "append",
                str(path),
                f"Worksheet '{sheet_name}' does not exist.",
            )

        if not isinstance(rows, (list, tuple)) or not rows:
            return _reject(
                "append",
                "rows must be a non-empty list of rows.",
            )

        width: Optional[int] = None

        for row_index, row in enumerate(rows):
            if not isinstance(row, (list, tuple)) or not row:
                return _reject(
                    "append",
                    f"Row {row_index} must be a non-empty list or tuple.",
                )

            width = width or len(row)

            if len(row) != width:
                return _reject(
                    "append",
                    "All appended rows must have the same number of cells.",
                )

            for column_index, value in enumerate(row):
                error = _validate_scalar(
                    value,
                    f"Cell ({row_index}, {column_index})",
                )

                if error:
                    return _reject("append", error)

        start_row = max(
            worksheet.max_row + 1,
            1,
        )

        if (
            start_row + len(rows) - 1 > _MAX_ROWS
            or width > _MAX_COLUMNS
        ):
            return _reject(
                "append",
                "Appended data exceeds worksheet limits.",
            )

        for row_offset, row in enumerate(rows):
            for column_offset, value in enumerate(row):
                worksheet.cell(
                    start_row + row_offset,
                    column_offset + 1,
                ).value = value

        expected_start = start_row
        expected_end = start_row + len(rows) - 1

        expected_bounds = (
            expected_start,
            1,
            expected_end,
            width,
        )

        expected_values = [
            [_json_value(value) for value in row]
            for row in rows
        ]

        def expected(verified: Any) -> Optional[str]:
            if sheet_name not in verified.sheetnames:
                return "Expected worksheet is missing after save."

            actual_values = _range_values(
                verified[sheet_name],
                expected_bounds,
            )

            if actual_values != expected_values:
                return (
                    "Appended rows do not match the requested data "
                    "after save."
                )

            return None

        save_error = _atomic_save(
            workbook,
            path,
            expected,
        )

        if save_error:
            return _error(
                "append",
                str(path),
                "Workbook append failed verification.",
                save_error,
            )

        return _result(
            True,
            "success",
            "append",
            str(path),
            "Rows appended successfully.",
            data={
                "sheet": sheet_name,
                "start_row": start_row,
                "rows_appended": len(rows),
            },
        )

    finally:
        workbook.close()


def _compute(
    path: Path,
    sheet_name: Any,
    reference: Any,
    computation: Any,
    overwrite: bool,
    destination: Any,
) -> Dict[str, Any]:
    workbook, failure = _load_existing(
        path,
        "compute",
    )

    if failure:
        return failure

    try:
        worksheet, failure = _sheet_or_error(
            workbook,
            sheet_name,
            "compute",
            str(path),
        )

        if failure:
            return failure

        bounds, error = _parse_reference(reference)

        if error:
            return _reject("compute", error)

        if (
            not isinstance(computation, str)
            or computation.lower() not in _COMPUTATIONS
        ):
            return _reject(
                "compute",
                "computation must be one of sum, average, min, max, or count.",
            )

        operation = computation.lower()

        raw_values = [
            value
            for row in _range_values(
                worksheet,
                bounds,
            )
            for value in row
        ]

        if operation == "count":
            value = sum(
                1
                for item in raw_values
                if item is not None
            )

        else:
            numeric_values = [
                item
                for item in raw_values
                if isinstance(item, (int, float))
                and not isinstance(item, bool)
            ]

            if not numeric_values:
                return _reject(
                    "compute",
                    "The selected range contains no numeric values.",
                )

            if operation == "sum":
                value = sum(numeric_values)

            elif operation == "average":
                value = sum(numeric_values) / len(numeric_values)

            elif operation == "min":
                value = min(numeric_values)

            else:
                value = max(numeric_values)

            if isinstance(value, float) and not math.isfinite(value):
                return _reject(
                    "compute",
                    "Computed result is not a finite number.",
                )

        if destination is None:
            return _result(
                True,
                "success",
                "compute",
                str(path),
                "Spreadsheet computation completed.",
                data={
                    "sheet": sheet_name,
                    "range": reference,
                    "computation": operation,
                    "value": value,
                },
            )

        if not isinstance(overwrite, bool) or not overwrite:
            return _reject(
                "compute",
                "Writing a computed result requires overwrite=True.",
            )

        destination_bounds, error = _parse_reference(destination)

        if error:
            return _reject("compute", error)

        if (
            destination_bounds[0] != destination_bounds[2]
            or destination_bounds[1] != destination_bounds[3]
        ):
            return _reject(
                "compute",
                "destination must be a single cell.",
            )

        destination_row = destination_bounds[0]
        destination_col = destination_bounds[1]

        worksheet.cell(
            destination_row,
            destination_col,
        ).value = value

        expected_destination = (
            destination_row,
            destination_col,
            destination_row,
            destination_col,
        )

        expected_value = _json_value(value)

        def expected(verified: Any) -> Optional[str]:
            if sheet_name not in verified.sheetnames:
                return "Expected worksheet is missing after save."

            actual = _range_values(
                verified[sheet_name],
                expected_destination,
            )

            if actual != [[expected_value]]:
                return (
                    "Computed result does not match the expected value "
                    "after save."
                )

            return None

        save_error = _atomic_save(
            workbook,
            path,
            expected,
        )

        if save_error:
            return _error(
                "compute",
                str(path),
                "Computed result write failed verification.",
                save_error,
            )

        return _result(
            True,
            "success",
            "compute",
            str(path),
            "Spreadsheet computation completed and was written.",
            data={
                "sheet": sheet_name,
                "range": reference,
                "computation": operation,
                "value": value,
                "destination": destination,
            },
        )

    finally:
        workbook.close()


def spreadsheet_tool(
    operation: Any,
    path: Any,
    sheet: Any = None,
    range: Any = None,
    data: Any = _MISSING,
    rows: Any = None,
    computation: Any = None,
    destination: Any = None,
    overwrite: bool = False,
    create: bool = False,
    create_sheet: bool = False,
    **kwargs: Any,
) -> Dict[str, Any]:
    """Perform one explicit, local XLSX operation.

    ``inspect`` needs only ``path``.
    ``read`` needs ``sheet`` and optionally ``range``.
    ``write`` needs ``sheet``, ``range``, and ``data``.
    ``append`` needs ``sheet`` and ``rows``.
    ``compute`` needs ``sheet``, ``range``, and ``computation``;
    pass ``destination`` to persist its result.
    """
    if kwargs:
        return _reject(
            operation if isinstance(operation, str) else None,
            "Unsupported spreadsheet options were supplied.",
        )

    if (
        not isinstance(operation, str)
        or operation.lower() not in _OPERATIONS
    ):
        return _reject(
            None,
            "operation must be one of inspect, read, write, append, or compute.",
        )

    operation = operation.lower()

    path_value, path_error = _validate_xlsx_path(path)

    if path_error:
        return _reject(
            operation,
            path_error,
        )

    try:
        if operation == "inspect":
            return _inspect(path_value)

        if operation == "read":
            return _read(
                path_value,
                sheet,
                range,
            )

        if operation == "write":
            if data is _MISSING:
                return _reject(
                    operation,
                    "data is required for write.",
                )

            return _write(
                path_value,
                sheet,
                range,
                data,
                overwrite,
                create,
                create_sheet,
            )

        if operation == "append":
            return _append(
                path_value,
                sheet,
                rows,
                overwrite,
                create,
                create_sheet,
            )

        return _compute(
            path_value,
            sheet,
            range,
            computation,
            overwrite,
            destination,
        )

    except PermissionError as exc:
        return _error(
            operation,
            str(path_value),
            "Permission denied.",
            type(exc).__name__,
        )

    except OSError as exc:
        return _error(
            operation,
            str(path_value),
            "Filesystem error.",
            type(exc).__name__,
        )

    except Exception as exc:  # pragma: no cover
        logger.exception(
            "Unexpected spreadsheet tool failure for operation '%s'.",
            operation,
        )

        return _error(
            operation,
            str(path_value),
            "Unexpected spreadsheet tool error.",
            type(exc).__name__,
        )


def tool(**kwargs: Any) -> Dict[str, Any]:
    """Executor-compatible keyword-only dispatch wrapper."""
    return spreadsheet_tool(**kwargs)