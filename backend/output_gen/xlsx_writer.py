"""
backend/output_gen/xlsx_writer.py

Local XLSX workbook writer for the Sovereign AI Workbench.

This module provides offline spreadsheet generation for Microsoft Excel .xlsx
files from structured content. It is a workspace-bound tool that enforces local
path security and requires no external APIs or network access.

Public API:
-----------
    write_xlsx(content, path, overwrite=False, **kwargs)

        Generate a .xlsx file from structured content.

        Args:
            content (dict): Structured workbook content with keys:
                - workbook (dict, optional): workbook metadata/title
                - sheets (list): list of sheet dicts

            path (str or Path): Output file path. Must resolve within
                                PROJECT_ROOT/data/ or PROJECT_ROOT/sandbox/.

            overwrite (bool): If False (default), reject if path already exists.
                              If True, replace existing file.

            **kwargs: Reserved for future extensions.

        Returns:
            dict: Stable result shape:
                {
                    "success": bool,
                    "path": str | None,
                    "format": "xlsx",
                    "message": str | None,
                    "error": str | None,
                }

Security & Constraints:
-----------------------
- Paths are validated and must remain within allowed workspace roots.
- Path traversal (../) and absolute paths outside allowed roots are rejected.
- No files are executed, evaluated, or interpreted as code.
- No environment variables or secrets are accessed.
- The implementation is completely offline and local.
"""

from __future__ import annotations

import logging
import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

__all__ = ["write_xlsx"]

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parents[2]

ALLOWED_ROOTS: List[Path] = [
    (PROJECT_ROOT / "data").resolve(),
    (PROJECT_ROOT / "sandbox").resolve(),
]

# Safety limits
_MAX_FILE_SIZE = 50_000_000  # 50 MB
_MAX_SHEETS = 100
_MAX_ROWS_PER_SHEET = 100_000
_MAX_COLUMNS_PER_SHEET = 100
_MAX_TEXT_LENGTH = 1_000_000


def _rejected_result(reason: str) -> Dict[str, Any]:
    """Build a stable rejected result dictionary."""
    return {
        "success": False,
        "path": None,
        "format": "xlsx",
        "message": None,
        "error": reason,
    }


def _error_result(path: Optional[str], reason: str) -> Dict[str, Any]:
    """Build a stable error result dictionary."""
    return {
        "success": False,
        "path": path,
        "format": "xlsx",
        "message": None,
        "error": reason,
    }


def _success_result(path: str) -> Dict[str, Any]:
    """Build a stable success result dictionary."""
    return {
        "success": True,
        "path": path,
        "format": "xlsx",
        "message": "XLSX generated successfully.",
        "error": None,
    }


def _resolve_within_allowed(path_value: Any) -> Optional[Path]:
    """Resolve a path and ensure it stays inside one of the allowed roots."""
    if path_value is None:
        return None
    if not isinstance(path_value, (str, Path)):
        return None
    try:
        candidate = Path(path_value).expanduser().resolve()
    except (TypeError, ValueError, OSError):
        return None
    for root in ALLOWED_ROOTS:
        try:
            candidate.relative_to(root)
        except ValueError:
            continue
        return candidate
    return None


def _validate_output_extension(path: Path) -> bool:
    """Validate that the output path has a .xlsx extension."""
    if not path.suffix:
        return False
    return path.suffix.lower() == ".xlsx"


def _valid_sheet_name(name: Any) -> bool:
    """Validate an Excel sheet name."""
    if not isinstance(name, str):
        return False
    if not name or not name.strip():
        return False
    if len(name) > 31:
        return False
    invalid_chars = set('[\\/:*?"<>|]')
    if any(ch in invalid_chars for ch in name):
        return False
    if name.endswith(".") or name.endswith(" "):
        return False
    return True


def _coerce_scalar(value: Any, field_name: str = "value") -> Any:
    """Validate supported scalar cell values and return them unchanged."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool, date, datetime)):
        if isinstance(value, str) and len(value) > _MAX_TEXT_LENGTH:
            raise ValueError(f"{field_name} exceeds maximum length ({_MAX_TEXT_LENGTH}).")
        return value
    raise ValueError(f"Unsupported value for {field_name}: {type(value).__name__}")


def _validate_text_length(value: Any, field_name: str) -> Optional[str]:
    """Reject text that exceeds the configured maximum length."""
    if value is None:
        return None
    if isinstance(value, str):
        if len(value) > _MAX_TEXT_LENGTH:
            return f"{field_name} exceeds maximum length ({_MAX_TEXT_LENGTH})."
    return None


def _validate_content(content: Any) -> Tuple[bool, Optional[str]]:
    """Validate the full workbook structure before generation begins."""
    if content is None:
        return False, "Content cannot be None."
    if not isinstance(content, dict):
        return False, "Content must be a dictionary."
    if len(content) == 0:
        return False, "Content dictionary cannot be empty."

    # Top-level workbook metadata block
    workbook_section = content.get("workbook")
    if workbook_section is not None:
        if not isinstance(workbook_section, dict):
            return False, "Workbook metadata must be a dictionary or None."
        title = workbook_section.get("title")
        if title is not None:
            if not isinstance(title, str):
                return False, "Workbook title must be a string or None."
            error = _validate_text_length(title, "Workbook title")
            if error:
                return False, error
        properties = workbook_section.get("properties")
        if properties is not None:
            if not isinstance(properties, dict):
                return False, "Workbook properties must be a dictionary or None."
            allowed_properties = {"creator", "subject", "description"}
            for key, value in properties.items():
                if not isinstance(key, str):
                    return False, "Workbook property names must be strings."
                if key not in allowed_properties:
                    return False, f"Unsupported workbook property: {key!r}."
                if value is not None:
                    if not isinstance(value, str):
                        return False, f"Workbook property '{key}' must be a string or None."
                    error = _validate_text_length(value, f"Workbook property '{key}'")
                    if error:
                        return False, error

    sheets = content.get("sheets")
    if sheets is None:
        return False, "Workbook must contain at least one sheet."
    if not isinstance(sheets, (list, tuple)):
        return False, "Sheets must be a list/tuple."
    if len(sheets) == 0:
        return False, "Workbook must contain at least one sheet."
    if len(sheets) > _MAX_SHEETS:
        return False, f"Too many sheets (max {_MAX_SHEETS})."

    seen_names = set()
    for sheet_index, sheet in enumerate(sheets):
        if not isinstance(sheet, dict):
            return False, f"Sheet {sheet_index} must be a dictionary."

        name = sheet.get("name")
        if name is None:
            name = f"Sheet{sheet_index + 1}"
        if not _valid_sheet_name(name):
            return False, f"Sheet {sheet_index} has an invalid name: {name!r}."
        if name in seen_names:
            return False, f"Duplicate sheet name: {name!r}."
        seen_names.add(name)

        headers = sheet.get("headers")
        if headers is not None:
            if not isinstance(headers, (list, tuple)):
                return False, f"Sheet '{name}' headers must be a list/tuple or None."
            if len(headers) == 0:
                return False, f"Sheet '{name}' must have at least one header column."
            if len(headers) > _MAX_COLUMNS_PER_SHEET:
                return False, f"Sheet '{name}' has too many columns (max {_MAX_COLUMNS_PER_SHEET})."
            for col_idx, header in enumerate(headers):
                if not isinstance(header, str):
                    return False, f"Header at column {col_idx} in sheet '{name}' must be a string."
                error = _validate_text_length(header, f"Header '{header}' in sheet '{name}'")
                if error:
                    return False, error

        rows = sheet.get("rows")
        if rows is None:
            rows = []
        if not isinstance(rows, (list, tuple)):
            return False, f"Sheet '{name}' rows must be a list/tuple or None."
        if len(rows) > _MAX_ROWS_PER_SHEET:
            return False, f"Sheet '{name}' has too many rows (max {_MAX_ROWS_PER_SHEET})."

        # Validate each row structure
        if headers is not None:
            expected_cols = len(headers)
            for row_index, row in enumerate(rows):
                if not isinstance(row, (list, tuple)):
                    return False, f"Row {row_index} in sheet '{name}' must be a list/tuple."
                if len(row) != expected_cols:
                    return False, (
                        f"Row {row_index} in sheet '{name}' has {len(row)} cells, "
                        f"expected {expected_cols}."
                    )
                for col_index, cell in enumerate(row):
                    try:
                        _coerce_scalar(cell, f"Cell ({row_index}, {col_index}) in sheet '{name}'")
                    except ValueError as exc:
                        return False, str(exc)

        # Validate any row entries even if no headers exist by ensuring they are all list/tuple and scalar values
        else:
            for row_index, row in enumerate(rows):
                if not isinstance(row, (list, tuple)):
                    return False, f"Row {row_index} in sheet '{name}' must be a list/tuple."
                if len(row) > _MAX_COLUMNS_PER_SHEET:
                    return False, f"Row {row_index} in sheet '{name}' exceeds max columns ({_MAX_COLUMNS_PER_SHEET})."
                for col_index, cell in enumerate(row):
                    try:
                        _coerce_scalar(cell, f"Cell ({row_index}, {col_index}) in sheet '{name}'")
                    except ValueError as exc:
                        return False, str(exc)

        # Validate top-level sheet text fields if present
        for field_name in ("title", "description"):
            if field_name in sheet:
                value = sheet[field_name]
                if value is not None and not isinstance(value, str):
                    return False, f"Sheet '{name}' {field_name} must be a string or None."
                error = _validate_text_length(value, f"Sheet '{name}' {field_name}")
                if error:
                    return False, error

    return True, None


def _write_sheet_data(ws, sheet: Dict[str, Any]) -> None:
    """Populate a worksheet from a validated sheet spec."""
    headers = sheet.get("headers")
    rows = sheet.get("rows") or []

    if headers is not None:
        ws.append([_coerce_scalar(cell, "header cell") for cell in headers])
        for row in rows:
            ws.append([_coerce_scalar(cell, "row cell") for cell in row])
        # Format header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
    else:
        for row in rows:
            ws.append([_coerce_scalar(cell, "row cell") for cell in row])

    # Basic column sizing based on content length
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            max_length = max(max_length, len(text))
        width = min(max(10, max_length + 2), 35)
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = width


def _apply_workbook_metadata(wb: Workbook, workbook_meta: Optional[Dict[str, Any]]) -> None:
    """Set workbook metadata on the workbook if provided."""
    if workbook_meta is None:
        return
    title = workbook_meta.get("title")
    if title is not None:
        wb.properties.title = title

    props = workbook_meta.get("properties")
    if props is not None:
        if "creator" in props:
            creator = props["creator"]
            if creator is not None:
                wb.properties.creator = _coerce_scalar(creator, "creator")
        if "subject" in props:
            subject = props["subject"]
            if subject is not None:
                wb.properties.subject = _coerce_scalar(subject, "subject")
        if "description" in props:
            description = props["description"]
            if description is not None:
                wb.properties.description = _coerce_scalar(description, "description")


def _build_workbook(content: Dict[str, Any]) -> Workbook:
    """Construct a Workbook from validated content."""
    workbook_meta = content.get("workbook")
    sheets = content.get("sheets")

    wb = Workbook()
    # Remove the default sheet to avoid creating an unintended blank first sheet
    if wb.sheetnames:
        default_name = wb.sheetnames[0]
        if default_name in wb.sheetnames:
            del wb[default_name]

    _apply_workbook_metadata(wb, workbook_meta)

    for index, sheet in enumerate(sheets):
        name = sheet.get("name")
        if name is None:
            name = f"Sheet{index + 1}"
        ws = wb.create_sheet(title=name)
        _write_sheet_data(ws, sheet)

    # If, for any reason, no sheet was created, add a default fallback sheet.
    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    return wb


def write_xlsx(
    content: Any,
    path: Any,
    overwrite: bool = False,
    **kwargs: Any,
) -> Dict[str, Any]:
    """
    Generate a Microsoft Excel .xlsx file from structured content.

    Args:
        content (dict): Structured workbook content.
        path (str or Path): Output file path. Must resolve within
                            PROJECT_ROOT/data/ or PROJECT_ROOT/sandbox/.
        overwrite (bool): If False (default), reject if file exists.
                          If True, replace existing file.
        **kwargs: Reserved for future use.

    Returns:
        dict: Result dictionary with keys:
            - success (bool): True if file was written successfully.
            - path (str | None): Resolved path if successful.
            - format (str): Always "xlsx".
            - message (str | None): Success message if successful.
            - error (str | None): Error message if unsuccessful.
    """

    resolved_path = _resolve_within_allowed(path)
    if resolved_path is None:
        return _rejected_result("Invalid 'path': must be inside data/ or sandbox/.")

    if not _validate_output_extension(resolved_path):
        return _rejected_result("Invalid file extension: must end with .xlsx")

    if resolved_path.exists():
        if not overwrite:
            return _rejected_result(
                f"File already exists: {resolved_path}. Set overwrite=True to replace."
            )
        if resolved_path.is_dir():
            return _error_result(str(resolved_path), "Path is a directory, not a file.")

    is_valid, validation_error = _validate_content(content)
    if not is_valid:
        return _rejected_result(validation_error or "Invalid content.")

    try:
        resolved_path.parent.mkdir(parents=True, exist_ok=True)
        wb = _build_workbook(content)

        temp_fd, temp_path = tempfile.mkstemp(
            suffix=".xlsx",
            dir=str(resolved_path.parent),
            prefix=".tmp_",
        )
        try:
            os.close(temp_fd)
            wb.save(temp_path)

            temp_path_obj = Path(temp_path)
            if not temp_path_obj.exists():
                return _error_result(str(resolved_path), "Temporary file was not created.")

            temp_file_size = temp_path_obj.stat().st_size
            if temp_file_size == 0:
                temp_path_obj.unlink(missing_ok=True)
                return _error_result(str(resolved_path), "Temporary file is empty (zero bytes).")

            if temp_file_size > _MAX_FILE_SIZE:
                temp_path_obj.unlink(missing_ok=True)
                return _rejected_result(
                    f"Generated file size ({temp_file_size} bytes) exceeds maximum limit ({_MAX_FILE_SIZE} bytes)."
                )

            temp_path_obj.replace(resolved_path)
        except Exception:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass
            raise

        if not resolved_path.exists():
            return _error_result(str(resolved_path), "File was not created.")

        file_size = resolved_path.stat().st_size
        if file_size == 0:
            return _error_result(str(resolved_path), "File is empty (zero bytes).")

        # Integrity check: ensure the produced XLSX is readable by openpyxl.
        # This catches corrupt/incomplete ZIP-based XLSX output before success
        # is reported to callers.
        from openpyxl import load_workbook
        try:
            with load_workbook(resolved_path, read_only=True, data_only=False) as verified_wb:
                if not verified_wb.sheetnames:
                    return _error_result(
                        str(resolved_path),
                        "Generated XLSX contains no worksheets.",
                    )
        except Exception as exc:
            try:
                resolved_path.unlink(missing_ok=True)
            except OSError:
                pass
            return _error_result(
                str(resolved_path),
                f"Generated XLSX failed integrity verification: {type(exc).__name__}",
            )

        return _success_result(str(resolved_path))

    except PermissionError as exc:
        return _error_result(str(resolved_path), f"Permission denied: {exc}")
    except OSError as exc:
        return _error_result(str(resolved_path), f"File system error: {exc}")
    except Exception as exc:
        logger.exception(f"Unexpected error generating XLSX: {exc}")
        return _error_result(str(resolved_path), f"Unexpected error: {type(exc).__name__}")
